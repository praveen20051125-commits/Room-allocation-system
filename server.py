# Useful commands
# Install dependencies:
#   pip install -r requirements.txt
#
# Run locally (development):
#   python server.py
#
# Run with Flask CLI:
#   Windows:  set FLASK_APP=server.py && flask run --host=0.0.0.0 --port=5000
#   Unix:     export FLASK_APP=server.py && flask run --host=0.0.0.0 --port=5000
#
# Production (Gunicorn):
#   gunicorn server:app --bind 0.0.0.0:$PORT --workers 2
#
# Render start command (use in Render service settings or Procfile):
#   web: gunicorn server:app --bind 0.0.0.0:$PORT --workers 2

import io
import os
from flask import Flask, render_template_string, request, redirect, url_for
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None  # openpyxl is optional at import time; upload will check and report an error if missing
import app as app_logic

app = Flask(__name__)
app.config["UPLOAD_FOLDER"] = os.path.join(os.path.dirname(__file__), "uploads")

HTML_TEMPLATE = """
<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Student Room Allocation</title>
  <style>
    body { font-family: Arial, sans-serif; margin: 20px; background: #f7f8fb; color: #1c1c1c; }
    .card { background: white; padding: 20px; border-radius: 12px; box-shadow: 0 4px 10px rgba(0,0,0,0.08); margin-bottom: 20px; }
    form { display: flex; gap: 10px; flex-wrap: wrap; align-items: center; }
    input, select, button { padding: 10px; border-radius: 8px; border: 1px solid #cfd8e3; }
    button { background: #1f6feb; color: white; cursor: pointer; }
    table { width: 100%; border-collapse: collapse; margin-top: 10px; }
    th, td { padding: 10px; border-bottom: 1px solid #e4e8ee; text-align: left; }
    .badge { display: inline-block; padding: 5px 10px; border-radius: 999px; background: #e8f2ff; color: #1f6feb; }

    /* Custom file input styles */
    .file-input-wrapper { display: flex; gap: 10px; align-items: center; }
    input[type="file"] { display: none; }
    .btn-file {
      display: inline-flex; align-items: center; gap: 8px;
      padding: 10px 14px; border-radius: 8px; border: 1px solid #cfd8e3; background: white; color: #1f6feb; cursor: pointer;
      box-shadow: 0 1px 2px rgba(0,0,0,0.03);
    }
    .btn-file:hover { background: #f0f6ff; }
    .file-name { color: #6b7280; font-size: 0.95rem; }
  </style>
</head>
<body>
  <div class="card">
    <h1>Student Room Allocation System</h1>
    <p>Upload monthly student Excel files. Each department can upload student details and the system assigns 25 students per room.</p>
  </div>

  <div class="card">
    <h2>Upload Student Excel</h2>
    <form method="post" action="/upload" enctype="multipart/form-data">
      <div class="file-input-wrapper">
        <input type="file" id="file" name="file" required>
        <label for="file" class="btn-file">Choose file</label>
        <span id="file-name" class="file-name">No file chosen</span>
      </div>
      <select name="department" required>
        <option value="CSE">CSE</option>
        <option value="ECE">ECE</option>
        <option value="EEE">EEE</option>
        <option value="MECH">MECH</option>
      </select>
      <button type="submit">Upload</button>
    </form>
  </div>

  <div class="card">
    <h2>Manual Removal</h2>
    <form method="post" action="/remove">
      <input type="text" name="register_number" placeholder="Register number" required>
      <button type="submit">Remove student</button>
    </form>
  </div>

  <div class="card">
    <h2>Room Summary</h2>
    <ul>
      {% for item in summary %}
        <li>Room {{ item.room_number }}: {{ item.total }} active students</li>
      {% endfor %}
    </ul>
  </div>

  <div class="card">
    <h2>Student List</h2>
    <table>
      <thead>
        <tr><th>Name</th><th>Register No</th><th>Department</th><th>Room</th><th>Status</th></tr>
      </thead>
      <tbody>
        {% for student in students %}
          <tr>
            <td>{{ student.name }}</td>
            <td>{{ student.register_number }}</td>
            <td>{{ student.department }}</td>
            <td>{{ student.room_number }}</td>
            <td><span class="badge">{% if student.is_active %}Active{% else %}Removed{% endif %}</span></td>
          </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>

  <script>
    (function(){
      const fileInput = document.getElementById('file');
      const fileName = document.getElementById('file-name');
      if (!fileInput) return;
      fileInput.addEventListener('change', function(){
        const name = this.files && this.files.length ? this.files[0].name : 'No file chosen';
        fileName.textContent = name;
      });
    })();
  </script>
</body>
</html>
"""


@app.route("/", methods=["GET"])
def index():
    app_logic.init_db()
    students = app_logic.get_students()
    summary = app_logic.get_room_summary()
    return render_template_string(HTML_TEMPLATE, students=students, summary=summary)


@app.route("/upload", methods=["POST"])
def upload():
    file = request.files.get("file")
    department = request.form.get("department", "")
    if not file or not department:
        return redirect(url_for("index"))

    if load_workbook is None:
        return (
            "Missing dependency: openpyxl is required to upload Excel files.\n"
            "Install it with: pip install openpyxl",
            500,
        )

    workbook = load_workbook(io.BytesIO(file.read()))
    sheet = workbook.active
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row):
            continue
        name = row[0] if len(row) > 0 else ""
        register_number = row[1] if len(row) > 1 else ""
        rows.append({"name": str(name), "register_number": str(register_number)})

    if rows:
        app_logic.import_students(rows, department=department)
    return redirect(url_for("index"))


@app.route("/remove", methods=["POST"])
def remove():
    register_number = request.form.get("register_number", "")
    if register_number:
        app_logic.remove_student(register_number)
    return redirect(url_for("index"))


if __name__ == "__main__":
    os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
    app.run(debug=True, host="0.0.0.0", port=5000)
